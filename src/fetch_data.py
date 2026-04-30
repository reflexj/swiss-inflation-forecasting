"""
fetch_data.py
-------------
Loads CPI, EUR/CHF, USD/CHF and IPI from locally saved files,
downloads oil price from FRED and converts it from USD to CHF.

Data sources:
    data/raw/cpi_bfs_raw.xlsx        <- BFS Excel (INDEX_m sheet)
    data/raw/eurchf_snb_raw.csv      <- SNB CSV (devkum, EUR, monthly)
    data/raw/usdchf_snb_raw.csv      <- SNB CSV (devkum, USD, monthly)
    data/raw/import_price_bfs_raw.xlsx <- BFS IPI Excel (INDEX_m sheet)
    FRED API                         <- Brent oil price (DCOILBRENTEU, USD/barrel)

Run:
    python src/fetch_data.py
"""

import os
import datetime
import requests
import pandas as pd
from fredapi import Fred
from dotenv import load_dotenv

# Paths
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW  = os.path.join(ROOT, "data", "raw")
os.makedirs(RAW, exist_ok=True)

load_dotenv(os.path.join(ROOT, ".env"))


def fetch_bfs_cpi(start: str = "1990-01-01") -> pd.DataFrame:
    """
    Reads the Swiss CPI index (LIK) from the manually downloaded BFS Excel file.
    File:   data/raw/cpi_bfs_raw.xlsx
    Sheet:  INDEX_m, Row 4 = header, Code '100_100' = Total index
    Returns DataFrame with columns: date, CPI_index, CPI_YoY
    """
    print("Reading CPI from local BFS Excel file...")

    xlsx_path = os.path.join(RAW, "cpi_bfs_raw.xlsx")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(
            f"BFS Excel file not found at:\n  {xlsx_path}\n"
            "Rename your downloaded BFS file to 'cpi_bfs_raw.xlsx' and place it in data/raw/"
        )

    df_raw = pd.read_excel(xlsx_path, sheet_name="INDEX_m", header=3, engine="openpyxl")
    total_row = df_raw[df_raw["Code"] == "100_100"].iloc[0]
    date_cols = [c for c in df_raw.columns if isinstance(c, datetime.datetime)]

    df = pd.DataFrame({
        "date": pd.to_datetime(date_cols),
        "CPI_index": [float(total_row[c]) for c in date_cols]
    })
    df = df.sort_values("date").reset_index(drop=True)
    df["CPI_YoY"] = df["CPI_index"].pct_change(12) * 100
    df = df[df["date"] >= start].reset_index(drop=True)

    path = os.path.join(RAW, "cpi_bfs.csv")
    df.to_csv(path, index=False)
    print(f"  Saved {len(df)} rows -> {path}")
    print(f"  Date range: {df['date'].min().date()} to {df['date'].max().date()}")
    return df


def fetch_snb_eurchf(start: str = "1999-01-01") -> pd.DataFrame:
    """
    Reads EUR/CHF from the manually downloaded SNB CSV file.
    File:   data/raw/eurchf_snb_raw.csv
    Format: semicolon-separated, 3 header rows, columns: Date;D0;D1;Value
    Note:   EUR/CHF only exists from March 1999 (Euro introduction)
    Returns DataFrame with columns: date, EUR_CHF
    """
    print("Reading EUR/CHF from local SNB CSV file...")

    csv_path = os.path.join(RAW, "eurchf_snb_raw.csv")
    if not os.path.exists(csv_path):
        raise FileNotFoundError(
            f"SNB CSV file not found at:\n  {csv_path}\n"
            "Rename your downloaded SNB file to 'eurchf_snb_raw.csv' and place it in data/raw/"
        )

    df = pd.read_csv(csv_path, sep=";", skiprows=3, encoding="utf-8-sig")
    df = df.rename(columns={"Date": "date", "Value": "EUR_CHF"})
    df["date"] = pd.to_datetime(df["date"] + "-01")
    df = df[["date", "EUR_CHF"]].sort_values("date").reset_index(drop=True)
    df = df[df["date"] >= start].reset_index(drop=True)

    path = os.path.join(RAW, "eurchf_snb.csv")
    df.to_csv(path, index=False)
    print(f"  Saved {len(df)} rows -> {path}")
    print(f"  Date range: {df['date'].min().date()} to {df['date'].max().date()}")
    return df


def fetch_snb_usdchf(start: str = "1990-01-01") -> pd.DataFrame:
    """
    Reads USD/CHF from the manually downloaded SNB CSV file.
    File:   data/raw/usdchf_snb_raw.csv
    Format: semicolon-separated, 3 header rows, columns: Date;D0;D1;Value
    Source: SNB data portal, 'devkum' table (monthly average), USD.
    Used to convert FRED Brent oil price (USD/barrel) into CHF/barrel.
    Returns DataFrame with columns: date, USD_CHF
    """
    print("Reading USD/CHF from local SNB CSV file...")

    csv_path = os.path.join(RAW, "usdchf_snb_raw.csv")
    if not os.path.exists(csv_path):
        raise FileNotFoundError(
            f"SNB CSV file not found at:\n  {csv_path}\n"
            "Download USD/CHF monthly averages from the SNB data portal "
            "(https://data.snb.ch -> devkum, filter currency = USD), "
            "rename the file to 'usdchf_snb_raw.csv' and place it in data/raw/"
        )

    df = pd.read_csv(csv_path, sep=";", skiprows=3, encoding="utf-8-sig")
    df = df.rename(columns={"Date": "date", "Value": "USD_CHF"})
    df["date"] = pd.to_datetime(df["date"] + "-01")
    df = df[["date", "USD_CHF"]].sort_values("date").reset_index(drop=True)
    df = df[df["date"] >= start].reset_index(drop=True)

    path = os.path.join(RAW, "usdchf_snb.csv")
    df.to_csv(path, index=False)
    print(f"  Saved {len(df)} rows -> {path}")
    print(f"  Date range: {df['date'].min().date()} to {df['date'].max().date()}")
    return df


def fetch_fred_oil(start: str = "1990-01-01") -> pd.DataFrame:
    """
    Downloads Brent crude oil price from FRED and converts it to CHF.
    Series: DCOILBRENTEU (daily, USD/barrel) -> resampled to monthly average
            -> multiplied by SNB monthly USD/CHF -> CHF/barrel.
    Requires FRED_API_KEY in .env and data/raw/usdchf_snb_raw.csv.
    Returns DataFrame with columns: date, oil_price  (CHF per barrel)
    """
    print("Fetching Brent Oil from FRED (USD) and converting to CHF...")

    api_key = os.getenv("FRED_API_KEY")
    if not api_key:
        raise EnvironmentError(
            "FRED_API_KEY not found. Copy .env.example to .env and add your key."
        )

    fred = Fred(api_key=api_key)
    raw = fred.get_series("DCOILBRENTEU", observation_start=start)

    df = raw.resample("MS").mean().reset_index()
    df.columns = ["date", "oil_price_usd"]
    df["date"] = pd.to_datetime(df["date"])

    # Load USD/CHF (already monthly, MS-aligned by fetch_snb_usdchf)
    usdchf_path = os.path.join(RAW, "usdchf_snb.csv")
    if not os.path.exists(usdchf_path):
        raise FileNotFoundError(
            f"USD/CHF CSV not found at:\n  {usdchf_path}\n"
            "Run fetch_snb_usdchf() first (it is called automatically in __main__)."
        )
    usdchf = pd.read_csv(usdchf_path, parse_dates=["date"])

    # Inner-merge on date: both series are monthly-start; any month without
    # a matching FX observation is dropped (consistent with preprocess.py).
    df = df.merge(usdchf, on="date", how="inner")
    df["oil_price"] = df["oil_price_usd"] * df["USD_CHF"]
    df = df[["date", "oil_price"]].sort_values("date").reset_index(drop=True)

    path = os.path.join(RAW, "oil_fred.csv")
    df.to_csv(path, index=False)
    print(f"  Saved {len(df)} rows -> {path}")
    print(f"  Date range: {df['date'].min().date()} to {df['date'].max().date()}")
    return df


def fetch_bfs_ipi(start: str = "1990-01-01") -> pd.DataFrame:
    """
    Reads the Swiss Import Price Index (IPI) from the manually downloaded BFS Excel file.
    File:   data/raw/import_price_bfs_raw.xlsx
    Sheet:  INDEX_m
    Structure: Row 8 onwards, col index 2=date, col index 6=IPI base 2010=100
    Returns DataFrame with columns: date, import_price
    """
    print("Reading Import Price Index from local BFS Excel file...")

    xlsx_path = os.path.join(RAW, "import_price_bfs_raw.xlsx")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(
            f"IPI Excel file not found at:\n  {xlsx_path}\n"
            "Rename your downloaded BFS IPI file to 'import_price_bfs_raw.xlsx' "
            "and place it in data/raw/"
        )

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["INDEX_m"]
    rows = list(ws.iter_rows(values_only=True))

    # Data starts at row 8 (index 7)
    # col index 2 = date, col index 6 = IPI base 2010=100
    data = []
    for row in rows[7:]:
        if len(row) > 6 and isinstance(row[2], datetime.datetime) and row[6] is not None:
            try:
                data.append({"date": row[2], "import_price": float(row[6])})
            except (TypeError, ValueError):
                pass

    df = pd.DataFrame(data)
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date").reset_index(drop=True)
    df = df[df["date"] >= start].reset_index(drop=True)

    path = os.path.join(RAW, "import_price_bfs.csv")
    df.to_csv(path, index=False)
    print(f"  Saved {len(df)} rows -> {path}")
    print(f"  Date range: {df['date'].min().date()} to {df['date'].max().date()}")
    return df


if __name__ == "__main__":
    fetch_bfs_cpi()
    fetch_snb_eurchf()
    fetch_snb_usdchf()   # must run before fetch_fred_oil (used for USD->CHF conversion)
    fetch_fred_oil()
    fetch_bfs_ipi()
    print("\nAll raw data loaded/downloaded successfully.")
    print("Next step: python src/preprocess.py")